"""Build public/bank-statement-template.xlsx.

Three-sheet template that helps Honeydue-leavers (and anyone who pulls
PDF/CSV statements from their TW bank instead of plugging an aggregator
into their account) transform a bank statement into Futari's universal
CSV format. See docs/superpowers/specs/csv-import-design.md for the
target schema.

Re-run after editing this file:

    python3 scripts/build-bank-statement-template.py

The output replaces public/bank-statement-template.xlsx in place.
"""

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------- shared styles ----------

INK = "1F2937"          # body text
MUTED = "6B7280"        # secondary text
ACCENT = "B8835E"       # warm brown (close to Futari --asset-color-house)
HEADER_BG = "FAF4EB"    # cream
ROW_ALT_BG = "FCFAF6"
BORDER = "E5DDD2"

font_title = Font(name="Helvetica Neue", size=15, bold=True, color=INK)
font_intro = Font(name="Helvetica Neue", size=11, color=INK)
font_header = Font(name="Helvetica Neue", size=11, bold=True, color=INK)
font_body = Font(name="Helvetica Neue", size=11, color=INK)
font_muted = Font(name="Helvetica Neue", size=10, color=MUTED, italic=True)

fill_header = PatternFill("solid", fgColor=HEADER_BG)
fill_alt = PatternFill("solid", fgColor=ROW_ALT_BG)

side = Side(style="thin", color=BORDER)
border = Border(left=side, right=side, top=side, bottom=side)

wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def set_col_widths(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def write_header_row(ws, row, values, comments=None):
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border
        cell.alignment = center
        if comments and comments[col_idx - 1]:
            cell.comment = Comment(comments[col_idx - 1], "Futari")
            cell.comment.width = 320
            cell.comment.height = 160


def write_body_row(ws, row, values, alt=False):
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = font_body
        cell.border = border
        cell.alignment = left
        if alt:
            cell.fill = fill_alt


# ---------- sheet 1: 轉換表 ----------

def build_sheet_main(wb):
    ws = wb.active
    ws.title = "轉換表"
    set_col_widths(ws, [14, 9, 11, 13, 12, 36])

    ws.merge_cells("A1:F1")
    intro = (
        "請把銀行對帳單的資料整理成下方欄位（或直接在這裡填）。"
        "整理完成後，將 A:F 欄複製貼到 Futari 匯入頁面，或將本工作表另存為 .csv（UTF-8 編碼）。\n"
        "Paste your bank-statement rows into the columns below, then copy A:F into"
        " Futari's import page, or export this sheet as a UTF-8 .csv."
    )
    cell = ws.cell(row=1, column=1, value=intro)
    cell.font = font_intro
    cell.fill = fill_header
    cell.alignment = wrap
    cell.border = border
    ws.row_dimensions[1].height = 78

    headers = ["日期", "類型", "金額", "類別", "成員", "備註"]
    comments = [
        (
            "格式：yyyy-MM-dd（也接受 yyyy/MM/dd、yyyyMMdd）。\n"
            "請用「記帳日期」欄位；若銀行只給「交易日期」也可接受。"
        ),
        (
            "可填值：支出 / 收入（也接受 expense / income，不分大小寫）。\n"
            "對帳單上的「借方／支出」→ 支出；「貸方／存入」→ 收入。\n"
            "帳戶之間的轉帳（如：活存 → 信用卡繳款）請整列刪除——Futari 不存轉帳。"
        ),
        (
            "正整數，不可加負號、千位逗號或小數。\n"
            "對帳單上的負號（−）或括號 (1,200) 都代表支出，請改用「類型」欄表達。\n"
            "USD 帳戶需 *100 後填整數（例：1.50 → 150）。"
        ),
        (
            "對應 Futari 內建分類：飲食 / 服飾 / 居住 / 交通 / 教育 / 娛樂 /"
            " 醫療 / 金融 / 其他。\n"
            "對帳單只有「摘要 / 說明」沒有分類；可參考「類別建議」工作表。\n"
            "留空 → 匯入時 fallback 為「其他」。"
        ),
        (
            "可填值：member_a / member_b。\n"
            "依持卡人或帳戶主人判斷實際付款人。\n"
            "留空 → 採用匯入頁設定的預設付款人。"
        ),
        (
            "自由文字，建議 ≤ 500 字。\n"
            "建議：把對帳單的「摘要 / 說明」加上你自己看得懂的補充。\n"
            "留空 → 匯入時自動以「匯入紀錄」帶入。"
        ),
    ]
    write_header_row(ws, 2, headers, comments)
    ws.row_dimensions[2].height = 28

    examples = [
        ["2025-04-12", "支出", 320, "飲食", "member_a", "7-11 巷口早餐"],
        ["2025-04-12", "支出", 1480, "居住", "member_b", "台水四月帳單"],
        ["2025-04-15", "收入", 65000, "金融", "member_a", "四月薪轉"],
    ]
    for i, row_values in enumerate(examples):
        write_body_row(ws, 3 + i, row_values, alt=(i % 2 == 1))
        ws.row_dimensions[3 + i].height = 22

    ws.merge_cells("A6:F6")
    hint = ws.cell(
        row=6,
        column=1,
        value=(
            "↑ 上面三列為範例，正式整理時請覆寫或刪除。"
            "  ↑ Sample rows above — overwrite or delete before importing."
        ),
    )
    hint.font = font_muted
    hint.alignment = wrap
    ws.row_dimensions[6].height = 22

    ws.freeze_panes = "A3"


# ---------- sheet 2: 常見銀行欄位對照 ----------

def build_sheet_field_map(wb):
    ws = wb.create_sheet("常見銀行欄位對照")
    set_col_widths(ws, [22, 14, 56])

    ws.merge_cells("A1:C1")
    cell = ws.cell(
        row=1,
        column=1,
        value=(
            "台灣常見銀行對帳單欄位 → Futari 欄位的對照。"
            "涵蓋台新、中信、國泰、玉山、富邦的活存對帳單與信用卡帳單。\n"
            "Map TW bank-statement columns (Taishin / CTBC / Cathay / E.Sun / Fubon)"
            " to the Futari columns in 轉換表."
        ),
    )
    cell.font = font_intro
    cell.fill = fill_header
    cell.alignment = wrap
    cell.border = border
    ws.row_dimensions[1].height = 56

    write_header_row(ws, 2, ["銀行常見欄位", "Futari 欄位", "備註"])
    ws.row_dimensions[2].height = 26

    rows = [
        (
            "交易日 / 交易日期 / 消費日期",
            "日期",
            "信用卡帳單通常給「消費日期」，活存對帳單給「交易日」；任一個都可以",
        ),
        (
            "記帳日 / 記帳日期 / 入帳日",
            "日期",
            "若同時有「交易日」和「記帳日」，建議用記帳日（實際扣款日）",
        ),
        (
            "交易金額 / 支出金額 / 借方金額",
            "金額",
            "取絕對值，並把「支出」標到類型欄；千位逗號要去掉",
        ),
        (
            "存入金額 / 貸方金額",
            "金額",
            "取絕對值，類型欄填「收入」",
        ),
        (
            "幣別 / 交易幣別",
            "—",
            "目前匯入僅支援 group 的 base 幣別；非 base 幣別交易請先換算成 base 後填入",
        ),
        (
            "摘要 / 交易說明 / 描述",
            "備註",
            "對帳單最有資訊量的欄位，建議直接帶入；可參考「類別建議」推類別",
        ),
        (
            "ATM / 跨行 / 通路代碼",
            "（併入備註）",
            "Futari 不存通路；想保留就接在備註後，例：「巷口 7-11｜跨行提款」",
        ),
        (
            "對方戶名 / 對方帳號",
            "（併入備註）",
            "轉帳對象資訊；併入備註保留語境",
        ),
        (
            "餘額",
            "—",
            "Futari 由全帳本重算，不需要也不會匯入餘額欄",
        ),
        (
            "銀行手續費 / 利息收入",
            "金額 + 類別＝金融",
            "獨立記成一筆，分類填「金融」",
        ),
        (
            "信用卡分期 / 分期付款明細",
            "—",
            "Futari MVP 不重建分期狀態；建議以「首期扣款日 + 總金額」記為一筆",
        ),
        (
            "信用卡退款 / 撤銷",
            "（沖銷原 row）",
            "Futari 不支援自動沖銷；建議整列刪除被退款的原始 row，避免雙重計算",
        ),
    ]
    for i, row_values in enumerate(rows):
        write_body_row(ws, 3 + i, row_values, alt=(i % 2 == 1))
        ws.row_dimensions[3 + i].height = 32

    foot_row = 3 + len(rows) + 1
    ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=3)
    foot = ws.cell(
        row=foot_row,
        column=1,
        value=(
            "提醒：跨帳戶轉帳（活存 → 信用卡繳款、儲蓄 → 投資）請整列刪除，"
            "Futari 不存轉帳概念，匯入會被算成支出 + 收入兩筆，平白衝高總額。\n"
            "Note: Drop inter-account transfer rows. Futari has no transfer concept;"
            " importing them double-counts both sides."
        ),
    )
    foot.font = font_muted
    foot.alignment = wrap
    ws.row_dimensions[foot_row].height = 56


# ---------- sheet 3: 類別建議 ----------

def build_sheet_category_hints(wb):
    ws = wb.create_sheet("類別建議")
    set_col_widths(ws, [26, 16, 42])

    ws.merge_cells("A1:C1")
    cell = ws.cell(
        row=1,
        column=1,
        value=(
            "對帳單只有「摘要 / 說明」欄位，沒有分類。以下是常見字串 → Futari 分類建議。"
            "用 Excel 篩選或公式批次套用，不必逐筆改。\n"
            "Bank statements only carry a description column; here's a starter mapping"
            " from common merchant strings to Futari categories. Filter or formula"
            " to apply in bulk."
        ),
    )
    cell.font = font_intro
    cell.fill = fill_header
    cell.alignment = wrap
    cell.border = border
    ws.row_dimensions[1].height = 64

    write_header_row(ws, 2, ["對帳單常見字串", "Futari 分類", "備註"])
    ws.row_dimensions[2].height = 26

    rows = [
        ("7-11 / 全家 / 萊爾富 / OK", "飲食", "便利商店多數是飲食或雜貨"),
        ("麥當勞 / 肯德基 / 摩斯 / 漢堡王", "飲食", "速食連鎖"),
        ("星巴克 / Louisa / cama", "飲食", "咖啡連鎖"),
        ("UberEats / FoodPanda / 外送平台", "飲食", "外送平台"),
        ("家樂福 / 大潤發 / 全聯 / Costco", "飲食", "生鮮量販"),
        ("UNIQLO / Zara / H&M / Net", "服飾", "服飾連鎖"),
        ("屈臣氏 / 康是美 / 寶雅 / 日藥本舖", "服飾", "美妝美容打理"),
        ("台水 / 台電 / 中油 / 瓦斯費", "居住", "水電瓦斯"),
        ("中華電信 / 遠傳 / 台灣大哥大", "居住", "電信費；通訊類也可放金融"),
        ("Netflix / Disney+ / Spotify / Apple", "娛樂", "訂閱串流"),
        ("特力屋 / IKEA / 宜得利", "居住", "家具家用"),
        ("捷運 / MRT / 悠遊卡儲值 / 一卡通", "交通", "大眾運輸"),
        ("台灣高鐵 / 台鐵 / 國光客運", "交通", "城際交通"),
        ("中油 / 台塑石油 / 加油站", "交通", "車輛油資"),
        ("ETC / 遠通電收 / 過路費", "交通", "高速公路費"),
        ("Uber / LINE TAXI / 計程車", "交通", "計程車"),
        ("PChome / 蝦皮 / momo / 樂天", "其他", "電商通常難分類；依品項自行細分"),
        ("國泰世華 / 中信 / 玉山 手續費", "金融", "手續費；分類填金融"),
        ("信用卡年費 / 循環利息", "金融", ""),
        ("健保署 / 全民健保 / 健保費", "醫療", ""),
        ("國泰人壽 / 富邦人壽 / 南山", "金融", "保費；分類填金融"),
        ("教育部 / 學費 / 補習班", "教育", ""),
        ("ATM 提款 / 跨行提款", "其他", "提款本身不是消費；建議拆成後續花費 row"),
        ("跨行轉帳 / 約定轉帳", "—", "若是繳信用卡 / 帳戶間調度，整列刪除"),
    ]
    for i, row_values in enumerate(rows):
        write_body_row(ws, 3 + i, row_values, alt=(i % 2 == 1))
        ws.row_dimensions[3 + i].height = 22

    ref_row = 3 + len(rows) + 1
    ws.merge_cells(start_row=ref_row, start_column=1, end_row=ref_row, end_column=3)
    ref = ws.cell(
        row=ref_row,
        column=1,
        value=(
            "Futari 九大分類：飲食 / 服飾 / 居住 / 交通 / 教育 / 娛樂 / 醫療 / 金融 / 其他。"
            "  Futari uses nine categories: dining / clothing / housing / transit /"
            " education / entertainment / health / financial / other."
        ),
    )
    ref.font = font_muted
    ref.alignment = wrap
    ws.row_dimensions[ref_row].height = 40


# ---------- main ----------

def main():
    wb = Workbook()
    build_sheet_main(wb)
    build_sheet_field_map(wb)
    build_sheet_category_hints(wb)

    out = "public/bank-statement-template.xlsx"
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
