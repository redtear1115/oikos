"""Build public/cwmoney-template.xlsx.

Three-sheet template that helps CWMoney users transform their export
into Futari's universal CSV format (see docs/superpowers/specs/
csv-import-design.md).

Re-run after editing this file:

    python3 scripts/build-cwmoney-template.py

The output replaces public/cwmoney-template.xlsx in place.
"""

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ---------- shared styles ----------

INK = "1F2937"          # body text
MUTED = "6B7280"        # secondary text
ACCENT = "B8835E"       # warm brown (close to Futari --asset-color-house)
HEADER_BG = "FAF4EB"    # cream
ROW_ALT_BG = "FCFAF6"
BORDER = "E5DDD2"

font_title = Font(name="Helvetica Neue", size=15, bold=True, color=INK)
font_intro = Font(name="Helvetica Neue", size=11, color=INK)
font_header = Font(name="Helvetica Neue", size=11, bold=True, color=INK)
font_body = Font(name="Helvetica Neue", size=11, color=INK)
font_muted = Font(name="Helvetica Neue", size=10, color=MUTED, italic=True)
font_en = Font(name="Helvetica Neue", size=10, color=MUTED)

fill_header = PatternFill("solid", fgColor=HEADER_BG)
fill_alt = PatternFill("solid", fgColor=ROW_ALT_BG)

side = Side(style="thin", color=BORDER)
border = Border(left=side, right=side, top=side, bottom=side)

wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)
left = Alignment(horizontal="left", vertical="center", wrap_text=True)


def set_col_widths(ws, widths):
    for idx, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = w


def write_header_row(ws, row, values, comments=None):
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = font_header
        cell.fill = fill_header
        cell.border = border
        cell.alignment = center
        if comments and comments[col_idx - 1]:
            cell.comment = Comment(comments[col_idx - 1], "Futari")
            cell.comment.width = 320
            cell.comment.height = 160


def write_body_row(ws, row, values, alt=False):
    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row, column=col_idx, value=value)
        cell.font = font_body
        cell.border = border
        cell.alignment = left
        if alt:
            cell.fill = fill_alt


# ---------- sheet 1: 轉換表 ----------

def build_sheet_main(wb):
    ws = wb.active
    ws.title = "轉換表"
    set_col_widths(ws, [14, 9, 11, 13, 12, 32])

    # Row 1 — title (merged) — long instruction
    ws.merge_cells("A1:F1")
    intro = (
        "請把 CWMoney 匯出的資料整理成下方欄位。整理完成後，將 A:F 欄複製貼到"
        " Futari 匯入頁面，或將本工作表另存為 .csv（UTF-8 編碼）。\n"
        "Paste your CWMoney rows into the columns below, then copy A:F into"
        " Futari's import page, or export this sheet as a UTF-8 .csv."
    )
    cell = ws.cell(row=1, column=1, value=intro)
    cell.font = font_intro
    cell.fill = fill_header
    cell.alignment = wrap
    cell.border = border
    ws.row_dimensions[1].height = 70

    # Row 2 — headers
    headers = ["日期", "類型", "金額", "類別", "成員", "備註"]
    comments = [
        (
            "格式：yyyy-MM-dd（也接受 yyyy/MM/dd、yyyyMMdd）。\n"
            "CWMoney 的 i_date 是 Unix 毫秒，可用 Excel 公式：\n"
            "=TEXT((i_date/86400000)+25569,\"yyyy-mm-dd\")"
        ),
        (
            "可填值：支出 / 收入（也接受 expense / income，不分大小寫）。\n"
            "CWMoney 內帳戶間轉帳（i_type=轉帳）不對應 Futari 概念，請手動移除。"
        ),
        (
            "正整數，不可加負號、千位逗號或小數。\n"
            "CWMoney 的 i_money 取絕對值後填入。\n"
            "USD 需 *100 後填整數（例：1.50 → 150）。"
        ),
        (
            "對應 Futari 內建分類：飲食 / 服飾 / 居住 / 交通 / 教育 / 娛樂 /"
            " 醫療 / 金融 / 其他。\n"
            "CWMoney 的 i_kind / i_kinds 對應表見「類別對照表」工作表。\n"
            "留空 → 匯入時 fallback 為「其他」。"
        ),
        (
            "可填值：member_a / member_b。\n"
            "依 CWMoney 的 i_account 對應到 Futari 兩位成員。\n"
            "留空 → 採用匯入頁設定的預設付款人。"
        ),
        (
            "自由文字，建議 ≤ 500 字。\n"
            "CWMoney 的 i_remarks 直接複製即可。\n"
            "留空 → 匯入時自動以「匯入紀錄」帶入。"
        ),
    ]
    write_header_row(ws, 2, headers, comments)
    ws.row_dimensions[2].height = 28

    # Rows 3–5 — example data
    examples = [
        ["2025-04-12", "支出", 320, "飲食", "member_a", "巷口早餐買兩份"],
        ["2025-04-12", "支出", 1480, "居住", "member_b", "本月水費"],
        ["2025-04-15", "收入", 65000, "金融", "member_a", "四月薪轉"],
    ]
    for i, row_values in enumerate(examples):
        write_body_row(ws, 3 + i, row_values, alt=(i % 2 == 1))
        ws.row_dimensions[3 + i].height = 22

    # Row 6 — gentle hint (merged), grey italic
    ws.merge_cells("A6:F6")
    hint = ws.cell(
        row=6,
        column=1,
        value=(
            "↑ 上面三列為範例，正式整理時請覆寫或刪除。"
            "  ↑ Sample rows above — overwrite or delete before importing."
        ),
    )
    hint.font = font_muted
    hint.alignment = wrap
    ws.row_dimensions[6].height = 22

    # Freeze top so headers stay visible while scrolling
    ws.freeze_panes = "A3"


# ---------- sheet 2: CWMoney 欄位對照 ----------

def build_sheet_field_map(wb):
    ws = wb.create_sheet("CWMoney 欄位對照")
    set_col_widths(ws, [20, 14, 56])

    ws.merge_cells("A1:C1")
    cell = ws.cell(
        row=1,
        column=1,
        value=(
            "把 CWMoney 匯出的欄位對應到「轉換表」需要的欄位。"
            "  Map CWMoney export columns to the Futari columns in 轉換表."
        ),
    )
    cell.font = font_intro
    cell.fill = fill_header
    cell.alignment = wrap
    cell.border = border
    ws.row_dimensions[1].height = 40

    write_header_row(ws, 2, ["CWMoney 欄位", "Futari 欄位", "備註"])
    ws.row_dimensions[2].height = 26

    rows = [
        (
            "i_date",
            "日期",
            "Unix 毫秒時間戳。Excel 公式：=TEXT((A2/86400000)+25569,\"yyyy-mm-dd\")",
        ),
        (
            "i_type",
            "類型",
            "CWMoney 用 0/1 或「支出/收入」字串標示；轉帳 row 請整列刪除（Futari 不存轉帳）",
        ),
        (
            "i_money",
            "金額",
            "取絕對值並去除小數位；USD 帳本請 *100 後填入整數",
        ),
        (
            "i_kind",
            "類別",
            "主類別字串，對應「類別對照表」中的 CWMoney 欄；無對應時填「其他」",
        ),
        (
            "i_kinds",
            "（併入備註）",
            "子類別字串。建議併到備註欄保留語境，例：「i_kind/i_kinds｜原備註」",
        ),
        (
            "i_remarks",
            "備註",
            "直接複製；超過 500 字請自行精簡",
        ),
        (
            "i_account",
            "成員",
            "依帳戶名稱判斷實際付款人。例：「我的信用卡」→ member_a；「太太現金」→ member_b",
        ),
        (
            "i_project",
            "（併入備註）",
            "Futari 沒有專案欄位；建議併到備註，例：「[旅遊-京都] 原備註」",
        ),
        (
            "i_imageurl / i_lat / i_lng / i_invoice",
            "—",
            "Futari 不匯入；請整欄忽略",
        ),
    ]
    for i, row_values in enumerate(rows):
        write_body_row(ws, 3 + i, row_values, alt=(i % 2 == 1))
        ws.row_dimensions[3 + i].height = 30

    # Footer note
    foot_row = 3 + len(rows) + 1
    ws.merge_cells(start_row=foot_row, start_column=1, end_row=foot_row, end_column=3)
    foot = ws.cell(
        row=foot_row,
        column=1,
        value=(
            "提醒：Futari 不重建 CWMoney 的轉帳與分帳狀態，僅保留每筆紀錄本身。"
            "  Note: Futari imports each row as-is and does not reconstruct CWMoney's"
            " account transfers or settlement state."
        ),
    )
    foot.font = font_muted
    foot.alignment = wrap
    ws.row_dimensions[foot_row].height = 40


# ---------- sheet 3: 類別對照表 ----------

def build_sheet_category_map(wb):
    ws = wb.create_sheet("類別對照表")
    set_col_widths(ws, [22, 16, 40])

    ws.merge_cells("A1:C1")
    cell = ws.cell(
        row=1,
        column=1,
        value=(
            "CWMoney 常見分類對應到 Futari 的九大分類。沒列到的請就近選擇，留空則歸入「其他」。"
            "  Common CWMoney categories mapped to Futari's nine categories. Pick the"
            " closest match for anything not listed; blank rows fall back to 其他."
        ),
    )
    cell.font = font_intro
    cell.fill = fill_header
    cell.alignment = wrap
    cell.border = border
    ws.row_dimensions[1].height = 56

    write_header_row(ws, 2, ["CWMoney 分類", "Futari 分類", "備註"])
    ws.row_dimensions[2].height = 26

    rows = [
        ("早餐 / 午餐 / 晚餐", "飲食", "三餐都歸飲食"),
        ("飲料 / 咖啡 / 點心", "飲食", "含手搖、零食"),
        ("外食 / 聚餐", "飲食", ""),
        ("食材 / 生鮮 / 雜貨", "飲食", "在家煮的食材"),
        ("服飾 / 鞋子 / 配件", "服飾", ""),
        ("美容 / 美髮 / 保養", "服飾", "Futari 將外觀打理併入服飾"),
        ("房租 / 房貸", "居住", ""),
        ("水費 / 電費 / 瓦斯 / 網路", "居住", "家用基礎開銷"),
        ("家具 / 家電 / 家用品", "居住", ""),
        ("交通費 / 大眾運輸", "交通", "捷運、公車、計程車"),
        ("加油 / 過路費 / 停車費", "交通", "車輛使用相關"),
        ("學費 / 補習 / 書籍", "教育", ""),
        ("線上課程 / 訂閱知識服務", "教育", ""),
        ("娛樂 / 電影 / 遊戲", "娛樂", ""),
        ("旅遊 / 住宿", "娛樂", "旅行相關（旅行帳本另在 Futari Trip 處理）"),
        ("運動 / 健身", "娛樂", ""),
        ("禮物 / 紅包", "娛樂", "送人的心意"),
        ("醫療 / 看診 / 掛號", "醫療", ""),
        ("藥品 / 保健食品", "醫療", ""),
        ("保險費", "金融", "壽險、車險、健康險"),
        ("投資 / 基金 / 股票手續費", "金融", ""),
        ("銀行手續費 / 利息", "金融", ""),
        ("捐款 / 公益", "其他", ""),
        ("雜支 / 其他", "其他", "找不到對應就放這"),
    ]
    for i, row_values in enumerate(rows):
        write_body_row(ws, 3 + i, row_values, alt=(i % 2 == 1))
        ws.row_dimensions[3 + i].height = 22

    # Footer — Futari category quick reference
    ref_row = 3 + len(rows) + 1
    ws.merge_cells(start_row=ref_row, start_column=1, end_row=ref_row, end_column=3)
    ref = ws.cell(
        row=ref_row,
        column=1,
        value=(
            "Futari 九大分類：飲食 / 服飾 / 居住 / 交通 / 教育 / 娛樂 / 醫療 / 金融 / 其他。"
            "  Futari uses nine categories: dining / clothing / housing / transit /"
            " education / entertainment / health / financial / other."
        ),
    )
    ref.font = font_muted
    ref.alignment = wrap
    ws.row_dimensions[ref_row].height = 40


# ---------- main ----------

def main():
    wb = Workbook()
    build_sheet_main(wb)
    build_sheet_field_map(wb)
    build_sheet_category_map(wb)

    out = "public/cwmoney-template.xlsx"
    wb.save(out)
    print(f"wrote {out}")


if __name__ == "__main__":
    main()
